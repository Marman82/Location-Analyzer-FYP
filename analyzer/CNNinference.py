# predict image
import torch
import torch.nn as nn
from torchvision.transforms import transforms
import numpy as np
from torch.autograd import Variable
from torchvision.models import squeezenet1_1
import torch.functional as F
from io import open
import os
from PIL import Image
import pathlib
import glob
# text recognition
import easyocr
import cv2
from matplotlib import pyplot as plt
# load excel
import pandas as pd
from openpyxl import load_workbook
# show map in plot
import plotly.express as px
import folium
from branca.element import Element
from flask import *

train_path = r"scene_detection/my_train/my_train"
pred_path = r"C:\Users\Mar\Desktop\Year4SemB\FYP\ocr"

# count excel row size
book = load_workbook('longlat.xlsx')
sheet = book['Sheet1']
row_count = sheet.max_row
# extract cell value
dataframe1 = pd.read_excel('longlat.xlsx')
exceli = 0
excelValue = ""
while exceli < row_count-1:
    value = dataframe1.iloc[exceli, 0]
    excelValue += value+","
    exceli += 1
    # print("Value: "+value)

excelValue = excelValue.split(',')
# print(excelValue)

# extract fodler name to categories
root = pathlib.Path(train_path)
classes = sorted([j.name.split('/')[-1] for j in root.iterdir()])
print(classes)

# CNN Network


class ConvNet(nn.Module):
    def __init__(self, num_classes=5):
        super(ConvNet, self).__init__()

        # Output size after convolution filter [(w-f+2P)/s]+1
        # w = width and height
        # f = kernel_size
        # p = padding
        # s = stride

        # Input shape=(batch size,channel size,width,height)
        self.conv1 = nn.Conv2d(
            in_channels=3, out_channels=12, kernel_size=3, stride=1, padding=1)
        # Shape = (256,12,150,150)
        self.bn1 = nn.BatchNorm2d(num_features=12)
        # Shape = (256,12,150,150)
        self.relu1 = nn.ReLU()
        # Shape = (256,12,150,150)

        self.pool = nn.MaxPool2d(kernel_size=2)
        # Reduce the image size by factor 2
        # Shape = (256,12,75,75)

        self.conv2 = nn.Conv2d(
            in_channels=12, out_channels=20, kernel_size=3, stride=1, padding=1)
        # Shape = (256,20,75,75)
        self.relu2 = nn.ReLU()
        # Shape = (256,20,75,75)

        self.conv3 = nn.Conv2d(
            in_channels=20, out_channels=32, kernel_size=3, stride=1, padding=1)
        # Shape = (256,32,75,75)
        self.bn3 = nn.BatchNorm2d(num_features=32)
        # Shape = (256,32,75,75)
        self.relu3 = nn.ReLU()
        # Shape = (256,32,75,75)

        self.fc = nn.Linear(in_features=32*150*150, out_features=num_classes)

        # Feed forward function

    def forward(self, input):
        # same as above
        output = self.conv1(input)
        output = self.bn1(output)
        output = self.relu1(output)

        output = self.pool(output)

        output = self.conv2(output)
        output = self.relu2(output)

        output = self.conv3(output)
        output = self.bn3(output)
        output = self.relu3(output)

        # Above output will be in matrix, wiht shape (256,32,75,75)
        output = output.view(-1, 32*150*150)
        # feed inside fully connected layer
        output = self.fc(output)

        return output


checkpoint = torch.load('best_checkpoint.model',
                        map_location=torch.device('cpu'))
model = ConvNet(num_classes=5)
model.load_state_dict(checkpoint)
model.eval()

transformer = transforms.Compose([
    transforms.Resize((300, 300)),
    transforms.ToTensor(),  # 0-255 to 0-1, numpy to tensor
    # 0-1 to [-1,1], formula x-mean/stand deviation
    transforms.Normalize([0.5, 0.5, 0.5], [0.5, 0.5, 0.5])
])

# prediction


def prediction(imgPath: str):
    image = Image.open(imgPath)
    image_tensor = transformer(image).float()
    image_tensor = image_tensor.unsqueeze_(0)

    if torch.cuda.is_available():
        image_tensor.cuda()

    input = Variable(image_tensor)
    output = model(input)
    index = output.data.numpy().argmax()
    pred = classes[index]
    return findCoordinates(pred)


def findCoordinates(predicted_category: str):
    image_path = glob.glob(pred_path+'/*.jpg')[0]
    pred_dict = {}

    long = 0
    lat = 0

    pred_dict[image_path[image_path.rfind('/')+1:]] = predicted_category
    print("Looping Categories for image: " +
          pred_dict[image_path[image_path.rfind('/')+1:]])

    for index, j in enumerate(excelValue):
        if j == pred_dict[image_path[image_path.rfind('/')+1:]]:
            long = dataframe1.iloc[index, 1]
            lat = dataframe1.iloc[index, 2]
            print("Long Lat: "+str(long), str(lat))

    if pred_dict[image_path[image_path.rfind('/')+1:]] == "Maclehose Trail":
        reader = easyocr.Reader(['en', 'ch_tra'], gpu=False, verbose=False)
        # reader can't read chinese file name
        result = reader.readtext('predict.jpg')
        # print(result)

        img = cv2.imread('predict.jpg')
        image_path = 0
        for detection in result:
            text = detection[1].replace(" ", "")  # [[width,height],'OCR text']
            # print(text)
            if ("KK" in text or "JK" in text or "HE" in text):
                print("Found Position: "+text)
                for index, j in enumerate(excelValue):
                    if j == text:
                        long = dataframe1.iloc[index, 1]
                        lat = dataframe1.iloc[index, 2]
                        print("Long Lat: "+str(long), str(lat))
            image_path += 1

    return [long, lat]
